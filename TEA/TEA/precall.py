# -*- coding: utf-8 -*-
"""
Created on Thu Aug 20 12:35:46 2020

@author: Melissa Gray

Calculating precision and recall
"""

#%% VARIABLES 7 IMPORTS

import pandas as pd
import numpy as np
import TEA.confusion_matrix as cm
from glob import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os.path

from scipy.spatial import distance
from scipy.cluster.hierarchy import linkage, dendrogram, set_link_color_palette
import matplotlib.pyplot as plt


#%% CLASS

class Precall():
    def __init__(self):
        return
    
    def _calculate_precision(self, tp, fp):     # (TP) / (TP+FP)
        precision = 0
        if (tp==0) and (fp==0):
            precision = "nan"
        else:
            precision = (tp) / (tp+fp)
        return precision
    def _calculate_recall(self, tp, fn):        # (TP) / (TP+FN)
        recall = 0
        if (tp==0) and (fn==0):
            recall = "nan"
        else:
            recall = (tp) / (tp+fn)
        return recall
    
    def calculate_precision_and_recall(self, matrices):   # [TP, FN, FP, TN]
        tax_id_precision = {}
        tax_id_recall = {}
        
        for name in matrices:
            for tax_id in matrices[name]:
                if (tax_id not in tax_id_precision):
                    tax_id_precision[tax_id] = {}
                tax_id_precision[tax_id][name] = self._calculate_precision(matrices[name][tax_id][0], matrices[name][tax_id][2])
                if (tax_id not in tax_id_recall):
                    tax_id_recall[tax_id] = {}
                tax_id_recall[tax_id][name] = self._calculate_recall(matrices[name][tax_id][0], matrices[name][tax_id][1])
        for name in matrices:
            for tax_id in tax_id_precision:
                if name not in tax_id_precision[tax_id]:
                    tax_id_precision[tax_id][name] = "nan"
                if name not in tax_id_recall[tax_id]:
                    tax_id_recall[tax_id][name] = "nan"
        return tax_id_precision, tax_id_recall
    
    
    

# For misc stuff
class Misc():
    def __init__(self):
        self.matrix_dict = {}
        self.matrix_tables = {}
        self.saved = {}
        self.cm_truth = ""
        self.input_path = ""
        return
    
    # GETTERS
    def get_matrx_dict(self):
        return self.matrix_dict
    def get_matrix_tables(self):
        return self.matrix_tables
    def get_saved(self):
        return self.saved
    def get_matrix_names(self):
        names_list = []
        for m in self.matrix_dict:
            names_list.append(m)
        return names_list
    def get_input_path(self):
        return self.input_path
    
    def set_truth(self, truth):
        self.cm_truth = truth
        return
    def set_input_path(self, ip):
        self.input_path = ip
        return
      
    def add_matrix(self, name, matrix):
        self.matrix_dict[name] = matrix
        self.saved[name] = False
        return print("\nAdded as matrix", name)
    
    def remove_matrix(self, name):
        if name in self.matrix_dict:
            del self.matrix_dict[name]
            if name in self.matrix_tables:
                del self.matrix_tables
        else:
            print("There is no matrix by the name \'{}\'".format(name))
        return
    
    def show_matrices(self, name=""):
        if name == "":      # print all
            print()
            for key in self.matrix_dict:
                print(key)
                #print("\t-", self.matrix_dict[key])
        elif name in self.matrix_dict:       # print one
            print(name)
            print("\t-", self.matrix_dict[name])
        else:
            print("There is no matrix by the name \'{}\'".format(name))
        return
    
    def create_table(self, name=""):
        Juice = cm.Confusion(os.path.join(self.input_path, self.cm_truth), "")
        if name == "":
            for name in self.matrix_dict:
                if name not in self.matrix_tables:
                    Juice.set_file_name(os.path.join(self.input_path, name))
                    self.matrix_tables[name] = Juice.create_matrix_table(Juice.reformat_matrix(Juice.add_other_info(self.matrix_dict[name])))
        elif name in self.matrix_dict:
            Juice.set_file_name(os.path.join(self.input_path, name))
            self.matrix_tables[name] = Juice.create_matrix_table(Juice.reformat_matrix(Juice.add_other_info(self.matrix_dict[name])))
        else:
            print("There is no matrix by the name \'{}\'".format(name))
        return
    
    def show_tables(self, name=''):
        if name == "":
            print()
            for key in self.matrix_tables:
                print(key)
                print(self.matrix_tables[key])
        elif name in self.matrix_tables:
            print(name)
            print(self.matrix_tables[name])
        else:
            print("There is no matrix table by the name \'{}\'".format(name))
            
        return
    
    def save_matrices_as_csv(self, file_path):
        Juice = cm.Confusion(os.path.join(self.input_path, self.cm_truth), "")
        for name in self.matrix_dict:
            if self.saved[name] == False:
                Juice.set_file_name(os.path.join(self.input_path, name))
                self.create_table(name)
                csv_name = os.path.join(file_path, self.cm_truth + " " + name)
                Juice.save_matrix_table(self.matrix_tables[name], csv_name)
        return
    
    def _get_name_and_rank(self):
        Chai = cm.comp.pp.Parser()
        truth_other = Chai.main(os.path.join(self.input_path, self.cm_truth), 1)
        tp = {}
        fn = {}
        fp = {}
        tn = {}
        skipped_tax_id = []
        
        for sample_num in truth_other:
            for m in self.matrix_dict:
                for tax_id in self.matrix_dict[m]:
                    if tax_id in truth_other[sample_num]:
                        tp[tax_id] = {}
                        tp[tax_id]["rank"] = truth_other[sample_num][tax_id][0]
                        tp[tax_id]["name"] = truth_other[sample_num][tax_id][1]
                        fn[tax_id] = {}
                        fn[tax_id]["rank"] = truth_other[sample_num][tax_id][0]
                        fn[tax_id]["name"] = truth_other[sample_num][tax_id][1]
                        fp[tax_id] = {}
                        fp[tax_id]["rank"] = truth_other[sample_num][tax_id][0]
                        fp[tax_id]["name"] = truth_other[sample_num][tax_id][1]
                        tn[tax_id] = {}
                        tn[tax_id]["rank"] = truth_other[sample_num][tax_id][0]
                        tn[tax_id]["name"] = truth_other[sample_num][tax_id][1]
                    else:
                        skipped_tax_id.append(tax_id)
        
        if len(skipped_tax_id) > 0:
            for name in self.matrix_dict:
                temp_other = Chai.main(os.path.join(self.input_path, name), 1)
                for sample_num in temp_other:
                    for tax_id in skipped_tax_id:
                        if tax_id in temp_other[sample_num]:
                            tp[tax_id] = {}
                            tp[tax_id]["rank"] = temp_other[sample_num][tax_id][0]
                            tp[tax_id]["name"] = temp_other[sample_num][tax_id][1]
                            fn[tax_id] = {}
                            fn[tax_id]["rank"] = temp_other[sample_num][tax_id][0]
                            fn[tax_id]["name"] = temp_other[sample_num][tax_id][1]
                            fp[tax_id] = {}
                            fp[tax_id]["rank"] = temp_other[sample_num][tax_id][0]
                            fp[tax_id]["name"] = temp_other[sample_num][tax_id][1]
                            tn[tax_id] = {}
                            tn[tax_id]["rank"] = temp_other[sample_num][tax_id][0]
                            tn[tax_id]["name"] = temp_other[sample_num][tax_id][1]
        return tp, fn, fp, tn
    
    def _get_true_positives(self, TP):
        for m in self.matrix_dict:
            for tax_id in self.matrix_dict[m]:
                if tax_id not in TP:
                    TP[tax_id] = {}
                TP[tax_id][m] = self.matrix_dict[m][tax_id][0]
        return TP
    def _get_false_negatives(self, FN):
        for m in self.matrix_dict:
            for tax_id in self.matrix_dict[m]:
                if tax_id not in FN:
                    FN[tax_id] = {}
                FN[tax_id][m] = self.matrix_dict[m][tax_id][1]
        return FN
    def _get_false_positives(self, FP):
        for m in self.matrix_dict:
            for tax_id in self.matrix_dict[m]:
                if tax_id not in FP:
                    FP[tax_id] = {}
                FP[tax_id][m] = self.matrix_dict[m][tax_id][2]
        return FP
    def _get_true_negatives(self, TN):
        for m in self.matrix_dict:
            for tax_id in self.matrix_dict[m]:
                if tax_id not in TN:
                    TN[tax_id] = {}
                TN[tax_id][m] = self.matrix_dict[m][tax_id][3]
        return TN
    
    def _organize_matrix(self):
        Juice = cm.Confusion(os.path.join(self.input_path, self.cm_truth), "")
        tp, fn, fp, tn = self._get_name_and_rank()
        
        names = self.get_matrix_names()
        TP = self._get_true_positives(tp)
        FN = self._get_false_negatives(fn)
        FP = self._get_false_positives(fp)
        TN = self._get_true_negatives(tn)
        
        all_tax_ids = set(TP.keys()) | set(FN.keys()) | set(FP.keys()) | set(TN.keys())
        matrix_sum = Juice.matrix_sum()
        
        for tax_id in all_tax_ids:
            for name in names:
                if (name not in TP[tax_id]) and (name not in FN[tax_id]) and (name not in FP[tax_id]) and (name not in TN[tax_id]):
                    TN[tax_id][name] = matrix_sum
                
                if (name not in TP[tax_id]):
                    TP[tax_id][name] = 0
                if name not in FN[tax_id]:
                    FN[tax_id][name] = 0
                if name not in FP[tax_id]:
                    FP[tax_id][name] = 0
        
        for tax_id in TP:
            total_TP = 0
            total_FN = 0
            total_FP = 0
            total_TN = 0
            for tool in TP[tax_id]:
                if (tool != "rank") and (tool != "name"):
                    total_TP += int(TP[tax_id][tool])
                    total_FN += FN[tax_id][tool]
                    total_FP += FP[tax_id][tool]
                    total_TN += TN[tax_id][tool]
            TP[tax_id]["Aggregate"] = total_TP
            FN[tax_id]["Aggregate"] = total_FN
            FP[tax_id]["Aggregate"] = total_FP
            TN[tax_id]["Aggregate"] = total_TN
        
        return TP, FN, FP, TN
    
    def organize_matrix(self):
        Calc = Precall()
        tp, fn, fp, tn = self._organize_matrix()
        precision, recall = Calc.calculate_precision_and_recall(self.matrix_dict)
        
        sorted_tp, sorted_fn, sorted_fp, sorted_tn, sorted_precision, sorted_recall = {}, {}, {}, {}, {}, {}
        sorted_tax_ids = sorted(set(tp.keys()) | set(fn.keys()) | set(fp.keys()) | set(tn.keys()))
        
        for key in sorted_tax_ids:
            sorted_tp[key] = tp[key]
            sorted_fn[key] = fn[key]
            sorted_fp[key] = fp[key]
            sorted_tn[key] = tn[key]
            sorted_precision[key] = precision[key]
            sorted_recall[key] = recall[key]
        
        
        tp_df = pd.DataFrame.from_dict(sorted_tp, orient="index")
        fn_df = pd.DataFrame.from_dict(sorted_fn, orient="index")
        fp_df = pd.DataFrame.from_dict(sorted_fp, orient="index")
        tn_df = pd.DataFrame.from_dict(sorted_tn, orient="index")
        precision_df = pd.DataFrame.from_dict(sorted_precision, orient="index")
        recall_df = pd.DataFrame.from_dict(sorted_recall, orient="index")
        return tp_df, fn_df, fp_df, tn_df, precision_df, recall_df
    
    def _write_col_title(self, path):
        workbook = load_workbook(path)
        for name in workbook.sheetnames:
            sheet = workbook[name]
            c = sheet["A1"]
            c.value = "Tax ID"
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')
        workbook.save(path)
        print("\nSaved as \'{}\'".format(path))
        return
    def save_as_excel(self, file_path, file_name):
        tp, fn, fp, tn, precision, recall = self.organize_matrix()
        
        excel_name = os.path.join(file_path, file_name + ".xlsx")
        
        with pd.ExcelWriter(excel_name) as writer:
            tp.to_excel(writer, sheet_name="True Positives")
            fn.to_excel(writer, sheet_name="False Negatives")
            fp.to_excel(writer, sheet_name="False Positives")
            tn.to_excel(writer, sheet_name="True Negatives")
            precision.to_excel(writer, sheet_name="Precision")
            recall.to_excel(writer, sheet_name="Recall")
        
        self._write_col_title(excel_name)
        return
    
    
    def main(self, gnd_truth, excel_name="TaxaPerformanceMetrics_byTool", gen_dir="", file_path="", csv="no"):
        gen_paths = glob(os.path.join(gen_dir, "*.profile"))
        self.input_path = gen_dir
        
        Juice = cm.Confusion(os.path.join(self.input_path, gnd_truth), "")
        self.set_truth(gnd_truth)
        
        for path in gen_paths:
            name = os.path.basename(path)
            if (name != gnd_truth) and (name not in self.matrix_dict):
                Juice.set_file_name(path)
                self.add_matrix(name, Juice.main("no"))
        
        if csv.lower() == "yes":
            self.save_matrices_as_csv(file_path)
        
        self.save_as_excel(file_path, excel_name)
        
        
        sheets = ["True Positives", "False Negatives", "False Positives", "True Negatives"]
        ranks = self.read_excel(sheets, os.path.join(file_path, excel_name + ".xlsx"))
        for sheet in sheets:
            for rank in ranks:
                self.create_dendrogram(sheet, rank, file_path, os.path.join(file_path, excel_name + ".xlsx"))
        print("The Dendrograms have been saved in {}.".format(file_path))
        return
    
    def read_excel(self,sheets, excel_path):
        ranks = []
        excel_df = pd.read_excel(excel_path, sheet_name=sheets)
        for rank in excel_df[sheets[0]].loc[:, 'rank']:
            if rank not in ranks:
                ranks.append(rank)
        return ranks
    
    def create_dendrogram(self, metric, rank, file_path, excel_path):
        df = pd.read_excel(excel_path, sheet_name=metric)
        #df = df[metric]
        
        to_remove = ['Tax ID', 'rank', 'name', 'Aggregate']
        cols = [col for col in df.columns if col not in to_remove]
        
        tmp_df = df[df['rank'] == rank]
        tool_array = []
        names = []
        for item in cols:
            res = tmp_df[item]
            if np.sum(res) == 0:
                continue
            tool_array.append(res.tolist())
            names.append(item.split('.')[0])
        tool_array = np.array(tool_array)
        
        if len(tool_array) > 1:
            bray_curt = distance.pdist(np.array(tool_array), 'braycurtis')
            
            link = linkage(bray_curt, 'average')
            set_link_color_palette(['y', 'c', 'g', 'm', 'r'])
            
            plt.figure(figsize=[6.4, 10.4], dpi=480)
            title = metric + ": " + rank + "-Dendrogram"
            plt.suptitle(title)
            den = dendrogram(link, orientation='right', labels=names)
            
            fn = title.replace(": ", "-")
            filename = fn.replace(" ", "_") + '.png'
            plt.savefig(os.path.join(file_path, filename), dpi=480, facecolor='#B4FFDC', transparent=False)
            
            plt.close()
            print("{} has been saved.".format(filename))
        #plt.show()
        
        # add arg to create subplot grouped by metric or rank (subplot='none'; 'metric'; 'rank')
        return


#%% MAIN

if __name__ == "__main__":
    '''
    Choco = Misc()
    
    Choco.main(["truth", "pred", "pred2", "pred3"], "C:\\Users\\milkg\\Documents\\", "Derp_Test", "yes")
    '''
    
